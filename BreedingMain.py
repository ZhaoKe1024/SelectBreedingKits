#!/user/zhao/miniconda3/envs/torch-0
# -*- coding: utf_8 -*-
# @Time : 2024/3/17 18:25
# @Author: ZhaoKe
# @File : BreedingMain.py
# @Software: PyCharm
import math
import random
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from inbreed_lib.graphfromtable import get_graph_from_data
from inbreed_lib.selector.GASelector import GASelector
from inbreed_lib.procedure.kinship_on_graph import Kinship
from inbreed_lib.func import IDGenerator


def run_main(file_path, gene_idx=None, result_file=None):
    idgenarator = IDGenerator(end_number=int(gene_idx)*1000, year=str(int(gene_idx) + 1))
    layergraph, vertex_layer, vertex_list, sheet_list = get_graph_from_data(file_path=file_path)
    # print(sheet_list)
    kinship = Kinship(graph=layergraph)
    sheet_list += [str(int(sheet_list[-1]) + 1)]
    # print("Sheet list:", sheet_list)
    year2idx = {}  # {"16": 0, "17": 1, "18": 2, "19": 3, "20": 4, "21": 5}
    for jdx, item in enumerate(sheet_list):
        year2idx[item] = jdx
    print("Sheet list:", year2idx)
    print("Load edges from", gene_idx)
    popus = []
    print("year idx", year2idx[gene_idx])
    # return
    if gene_idx == sheet_list[-1]:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            print(vertex_list[item].name)
            popus.append(vertex_list[item])
    else:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            # print(vertex_list[item].name)
            popus.append(vertex_list[item])
            # popus.append(Poultry(fi=f_i, wi=wi, fa_i=fa_i, ma_i=ma_i, sex=0, inbreedc=0.))
    # return

    # print(kinship.calc_inbreed_coef(p="14774"))
    # kinship.print_edges()
    # kinship.print_layer()
    # kinship.print_parents()
    # kinship.print_all_poultry()
    # print(kinship.calc_kinship_corr(p1="14761", p2="14766"))
    random.shuffle(popus)
    male_rate = 1. / 11.
    male_num = math.ceil(male_rate * len(popus))
    female_num = len(popus) - male_num
    for i in range(male_num):
        vertex_list[i].gender = 1
        popus[i].sex = 1
    name2idx = dict()
    for i, p in enumerate(popus):
        name2idx[p.name] = i
    # -------------Kinship read and build-------------

    kinship_matrix = np.zeros((male_num, female_num))
    for i in range(male_num):
        for j in range(female_num):
            kinship_matrix[i][j] = kinship.calc_kinship_corr(p1=popus[i].name, p2=popus[male_num + j].name)
    # print(kinship_matrix)
    # print("max min")
    # print(np.max(kinship_matrix), np.min(kinship_matrix))
    # print(np.sum(kinship_matrix))

    # ===============================Algorithm==================================
    # ====================Here is vanilla Genetic Algorithm=====================

    GAS = GASelector(popus=popus, kinship_matrix=kinship_matrix, male_idxs=list(range(male_num)),
                     female_idxs=list(range(male_num, len(popus))))
    best_solution = GAS.scheduler()

    pre_pos = best_solution.vector_male[0]
    cur_female = best_solution.vector_female[0]
    # print("========----------育种方案----------==========")
    # print("(家系号，雄性个体编号, 雌性个体编号)]")
    idx = 1
    fout = open(result_file, 'w', encoding="utf_8")
    fout.write(
        "配种方案,家系号,公鸡号,母鸡号,亲缘相关系数,出雏,批次,翅号,公鸡号,母鸡号,{}年家系号,性别\n".format(gene_idx))
    # print("配种方案", "家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号", "{}年家系号".format(gene_idx))
    year, mi, fi = "21", 1, 1
    # print(
    #     f"{popus[pre_pos].family_id},{pre_pos}:[({popus[cur_female].family_id},{male_num + cur_female})",
    #     end=', ')
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    sex_id = idgenarator.get_rand_gender()
    child_id = idgenarator.get_new_id()
    fout.write(","  # col 1
               + tmp_fid + ","  # col 2 家系号
               + popus[pre_pos].name + ","  # col 3 公号
               + popus[male_num + cur_female].name + ","  # col 4 母号
               + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
               + child_id + ','
               + popus[pre_pos].name + ","  # col 9 公号
               + popus[male_num + cur_female].name + ","  # col 10 母号
               + tmp_fid + ","
               + sex_id + '\n')  # 11 家系号
    fi += 1
    # threshold = 0.5
    res_data = []
    while idx < len(best_solution):
        cur_male = best_solution.vector_male[idx]
        cur_male_name = popus[cur_male].name
        cur_female = best_solution.vector_female[idx]
        cur_female_name = popus[male_num + cur_female].name
        if cur_male != pre_pos:
            mi += 1
        fi += 1
        # <<<<<<< Updated upstream
        # fout.write(get_familyid(year, mi,
        #                         fi) + "," + cur_male_name + "," + cur_female_name + "," + f"{kinship_matrix[cur_male, cur_female]:.5f}" + '\n')
        # =======
        ibc = kinship_matrix[cur_male, cur_female]
        tmp_fid = idgenarator.get_family_id(y="", m=mi)
        sex_id = idgenarator.get_rand_gender()
        child_id = idgenarator.get_new_id()
        fout.write(","  # col 1
                   + tmp_fid + ","  # col 2 家系号
                   + popus[pre_pos].name + ","  # col 3 公号
                   + popus[male_num + cur_female].name + ","  # col 4 母号
                   + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
                   + child_id + ','
                   + popus[pre_pos].name + ","  # col 9 公号
                   + popus[male_num + cur_female].name + ","  # col 10 母号
                   + tmp_fid + ","
                   + sex_id + '\n')
        # print(tmp_fid + "," + cur_male_name + "," + cur_female_name + "," + f"{ibc:.5f}")
        res_data.append([tmp_fid, cur_male_name, cur_female_name, f"{ibc:.4f}", None, None, child_id, cur_male_name,
                         cur_female_name, tmp_fid, sex_id])
        # >>>>>>> Stashed changes
        pre_pos = cur_male
        idx += 1
    # print("]")
    fout.write('\n')
    fout.close()
    print(f"generate finished gene {gene_idx}")
    return res_data


def run_main_eq(file_path, gene_idx=None, result_file=None):
    """
    等量留种方式
    :param file_path:
    :param gene_idx:
    :param result_file:
    :return:
    """
    idgenarator = IDGenerator(end_number=int(gene_idx)*1000, year=str(int(gene_idx) + 1))
    layergraph, vertex_layer, vertex_list, sheet_list = get_graph_from_data(file_path=file_path)
    # print(sheet_list)
    kinship = Kinship(graph=layergraph)
    sheet_list += [str(int(sheet_list[-1]) + 1)]
    # print("Sheet list:", sheet_list)
    year2idx = {}  # {"16": 0, "17": 1, "18": 2, "19": 3, "20": 4, "21": 5}
    for jdx, item in enumerate(sheet_list):
        year2idx[item] = jdx
    print("Sheet list:", year2idx)
    print("Load edges from", gene_idx)
    popus = []
    print("year idx", year2idx[gene_idx])
    # return
    if gene_idx == sheet_list[-1]:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            print(vertex_list[item].name)
            popus.append(vertex_list[item])
    else:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            # print(vertex_list[item].name)
            popus.append(vertex_list[item])
            # popus.append(Poultry(fi=f_i, wi=wi, fa_i=fa_i, ma_i=ma_i, sex=0, inbreedc=0.))
    # return

    # print(kinship.calc_inbreed_coef(p="14774"))
    # kinship.print_edges()
    # kinship.print_layer()
    # kinship.print_parents()
    # kinship.print_all_poultry()
    # print(kinship.calc_kinship_corr(p1="14761", p2="14766"))
    random.shuffle(popus)
    male_rate = 1. / 11.
    male_num = math.ceil(male_rate * len(popus))
    female_num = len(popus) - male_num
    for i in range(male_num):
        vertex_list[i].gender = 1
        popus[i].sex = 1
    name2idx = dict()
    for i, p in enumerate(popus):
        name2idx[p.name] = i
    # -------------Kinship read and build-------------

    kinship_matrix = np.zeros((male_num, female_num))
    for i in range(male_num):
        for j in range(female_num):
            kinship_matrix[i][j] = kinship.calc_kinship_corr(p1=popus[i].name, p2=popus[male_num + j].name)
    # print(kinship_matrix)
    # print("max min")
    # print(np.max(kinship_matrix), np.min(kinship_matrix))
    # print(np.sum(kinship_matrix))

    # ===============================Algorithm==================================
    # ====================Here is vanilla Genetic Algorithm=====================

    GAS = GASelector(popus=popus, kinship_matrix=kinship_matrix, male_idxs=list(range(male_num)),
                     female_idxs=list(range(male_num, len(popus))))
    best_solution = GAS.scheduler()

    # ===========================找出最佳雌性，来生育最佳雄性，等数留种=====================
    female_list = []
    tmp_female_idx = 0
    tmp_female_value = 999.
    i, j = 0, 0
    tmp_male_idx = None
    cur_male_idx = best_solution.vector_male[0]
    while i < len(best_solution):
        tmp_male_idx = best_solution.vector_male[0]
        if cur_male_idx == tmp_male_idx:
            tmp_female_idx = best_solution.vector_female[0]
            if kinship_matrix[tmp_male_idx][tmp_female_idx] < tmp_female_value:
                tmp_female_value = kinship_matrix[tmp_male_idx][tmp_female_idx]
        else:
            female_list.append(tmp_female_idx)
        i += 1
    female_list.append(tmp_female_idx)
    # =============================================================================

    pre_pos = best_solution.vector_male[0]
    cur_female = best_solution.vector_female[0]
    # print("========----------育种方案----------==========")
    # print("(家系号，雄性个体编号, 雌性个体编号)]")
    idx = 1
    fout = open(result_file, 'w', encoding="utf_8")
    fout.write(
        "配种方案,家系号,公鸡号,母鸡号,亲缘相关系数,出雏,批次,翅号,公鸡号,母鸡号,{}年家系号,性别\n".format(gene_idx))
    # print("配种方案", "家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号", "{}年家系号".format(gene_idx))
    year, mi, fi = "21", 1, 1
    # print(
    #     f"{popus[pre_pos].family_id},{pre_pos}:[({popus[cur_female].family_id},{male_num + cur_female})",
    #     end=', ')
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    sex_id = "0"
    child_id = idgenarator.get_new_id()
    fout.write(","  # col 1
               + tmp_fid + ","  # col 2 家系号
               + popus[pre_pos].name + ","  # col 3 公号
               + popus[male_num + cur_female].name + ","  # col 4 母号
               + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
               + child_id + ','
               + popus[pre_pos].name + ","  # col 9 公号
               + popus[male_num + cur_female].name + ","  # col 10 母号
               + tmp_fid + ","
               + sex_id + '\n')  # 11 家系号
    fi += 1
    # threshold = 0.5
    res_data = []
    while idx < len(best_solution):
        cur_male = best_solution.vector_male[idx]
        cur_male_name = popus[cur_male].name
        cur_female = best_solution.vector_female[idx]
        cur_female_name = popus[male_num + cur_female].name
        if cur_male != pre_pos:
            mi += 1
        fi += 1
        # <<<<<<< Updated upstream
        # fout.write(get_familyid(year, mi,
        #                         fi) + "," + cur_male_name + "," + cur_female_name + "," + f"{kinship_matrix[cur_male, cur_female]:.5f}" + '\n')
        # =======
        ibc = kinship_matrix[cur_male, cur_female]
        tmp_fid = idgenarator.get_family_id(y="", m=mi)
        sex_id = idgenarator.get_rand_gender()
        child_id = idgenarator.get_new_id()  # 每只雌性家禽都留下一只后代
        fout.write(","  # col 1
                   + tmp_fid + ","  # col 2 家系号
                   + popus[pre_pos].name + ","  # col 3 公号
                   + popus[male_num + cur_female].name + ","  # col 4 母号
                   + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
                   + child_id + ','
                   + popus[pre_pos].name + ","  # col 9 公号
                   + popus[male_num + cur_female].name + ","  # col 10 母号
                   + tmp_fid + ","
                   + sex_id + '\n')
        # print(tmp_fid + "," + cur_male_name + "," + cur_female_name + "," + f"{ibc:.5f}")
        res_data.append([tmp_fid, cur_male_name, cur_female_name, f"{ibc:.4f}", None, None, child_id, cur_male_name,
                         cur_female_name, tmp_fid, sex_id])
        # >>>>>>> Stashed changes
        pre_pos = cur_male
        idx += 1
    # print("]")
    fout.write('\n')
    fout.close()
    print(f"generate finished gene {gene_idx}")
    return res_data


def run_n_generations(input_start, input_end, last_n):
    """

    :param input_start: read from sheet input_start
    :param input_end: read end at sheet input_end
    :param last_n: inbreeding last n year
    :return:
    """
    input_file = "../temp_files/历代配种方案及出雏对照2021_带性别.xlsx"
    name_tmplate = "../temp_files/历代配种方案及出雏对照2021_带性别{}.xlsx"
    cur_file = input_file
    res_data = None
    for f_year in range(2017, 2018):
        # run_main(calc.file_to_analyze, gene_idx=str("f_year"))

        res_data = run_main(file_path=cur_file, gene_idx=str(f_year), result_file="./test.csv")

        book = load_workbook(cur_file)
        writer = pd.ExcelWriter(name_tmplate.format(f_year), engine='openpyxl')
        writer.book = book
        df1 = pd.DataFrame(np.array(res_data))
        df1.columns = ["家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号",
                       "{}年家系号".format(str(f_year)[-2:]), "性别"]
        df1.to_excel(writer, str(f_year))  # first是第一张工作表名称
        writer.save()
        writer.close()
        cur_file = name_tmplate.format(f_year)
    return res_data


if __name__ == '__main__':
    # for i in [17, 18, 19, 20]:
    #     run_main(gene_idx=str(i))
    input_file = "../temp_files/历代配种方案及出雏对照2021_带性别.xlsx"
    name_tmplate = "../temp_files/历代配种方案及出雏对照2021_带性别{}.xlsx"
    cur_file = input_file
    for f_year in range(2017, 2018):
        # run_main(calc.file_to_analyze, gene_idx=str("f_year"))

        df1 = run_main(file_path=cur_file, gene_idx=str(f_year), result_file="./test.csv")

        book = load_workbook(cur_file)
        writer = pd.ExcelWriter(name_tmplate.format(f_year), engine='openpyxl')
        writer.book = book
        df1 = pd.DataFrame(np.array(df1))
        df1.columns = ["家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号",
                       "{}年家系号".format(str(f_year)[-2:]), "性别"]
        df1.to_excel(writer, str(f_year))  # first是第一张工作表名称
        writer.save()
        writer.close()
        cur_file = name_tmplate.format(f_year)
