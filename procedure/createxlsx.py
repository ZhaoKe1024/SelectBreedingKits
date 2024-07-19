# -*- coding: utf-8 -*-
# @Author : ZhaoKe
# @Time : 2024-07-20 0:23
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment

def create_template(gene_y="21"):
    xlsxbook = Workbook()

    # # --------
    # workSheet = xlsxbook.get_active_sheet()
    # workSheet.title = gene_y

    # -----
    workSheet = xlsxbook.create_sheet(gene_y, 5)
    workSheet.merge_cells('A1:A12')
    workSheet.cell(1,1).value = '人工输精表'
    workSheet.cell(2,1).value = "品种："
    workSheet.cell(2, 2).value = "青壳"
    workSheet.cell(2, 6).value = "鸡舍："

    workSheet.cell(2, 2).value = "青壳"
