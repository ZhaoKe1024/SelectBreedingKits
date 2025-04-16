#!/user/zhao/miniconda3/envs/torch-0
# -*- coding: utf_8 -*-
# @Time : 2024/11/1 15:15
# @Author: ZhaoKe
# @File : simplega.py
# @Software: PyCharm
# Equal Seed Reservation
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from inbreed_lib.graphfromtable import get_graph_from_data
from inbreed_lib.selector.entities import Vertex, calculate_fitness
from inbreed_lib.selector.GASelector import GASelector
from inbreed_lib.procedure.kinship_on_graph import Kinship
from inbreed_lib.func import IDGenerator


def run_main_without_graph(file_path="./kinship330.csv", gene_idx=2018, result_file=None):
    df = pd.read_csv(file_path, delimiter=',', header=0, index_col=None)
    popus_idxs = list(df.iloc[:, 0]) + list(df.columns[1:])
    popus = []
    for idx, name in enumerate(popus_idxs):
        popus.append(Vertex(idx, name=name))
    kinship_matrix = np.array(df.iloc[:, 1:])
    # print(kinship_matrix.shape)
    male_num = 30
    # ===============================Algorithm==================================
    # ====================Here is vanilla Genetic Algorithm=====================
    GAS = GASelector(popus=popus, kinship_matrix=kinship_matrix, male_idxs=list(range(male_num)),
                     female_idxs=list(range(male_num, len(popus))), num_popu=50, num_iter=20, mode="v1")
    best_solution = GAS.scheduler()

    best_solution.vector_female = list(range(300))
    best_solution.fitness_value = calculate_fitness(best_solution, kinship_matrix, mode="v2")

    print(best_solution.vector_male)
    print(best_solution.vector_female)
    print(best_solution.fitness_value)
    # ===========================找出最佳雌性，来生育最佳雄性，等数留种=====================
    female_list = []
    tmp_female_idx = 0
    i, j = 0, 0
    bst_female_value = 999.
    bst_female_idx = None
    cur_male_idx = best_solution.vector_male[0]
    while i < len(best_solution):
        tmp_male_idx = best_solution.vector_male[i]
        if cur_male_idx == tmp_male_idx:
            tmp_female_idx = best_solution.vector_female[i]
            if kinship_matrix[tmp_male_idx][tmp_female_idx] < bst_female_value:
                bst_female_value = kinship_matrix[tmp_male_idx][tmp_female_idx]
                bst_female_idx = tmp_female_idx
        else:
            female_list.append(bst_female_idx)
            bst_female_value = 999.
            cur_male_idx = tmp_male_idx
        i += 1
    female_list.append(bst_female_idx)
    print("best females:", len(female_list))
    print(female_list)

    # =============================================================================
    # print("========----------育种方案----------==========")
    # print("(家系号，雄性个体编号, 雌性个体编号)]")
    idgenarator = IDGenerator(end_number=int(gene_idx) * 1000, year=str(int(gene_idx) + 1))

    fout = open(result_file, 'w', encoding="utf_8")
    fout.write(
        "配种方案,家系号,公鸡号,母鸡号,亲缘相关系数,出雏,批次,翅号,公鸡号,母鸡号,{}年家系号,性别\n".format(gene_idx))
    # print("配种方案", "家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号", "{}年家系号".format(gene_idx))
    year, mi, fi = "24", 0, 0
    best_idx = 0
    # print(
    #     f"{popus[pre_pos].family_id},{pre_pos}:[({popus[cur_female].family_id},{male_num + cur_female})",
    #     end=', ')

    pre_pos = best_solution.vector_male[female_list[best_idx]]
    cur_female = best_solution.vector_female[female_list[best_idx]]
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    sex_id = "1"
    child_id = idgenarator.get_new_id()
    fout.write(","  # col 1
               + tmp_fid + ","  # col 2 家系号
               + popus[pre_pos].name + ","  # col 3 公号
               + popus[male_num + cur_female].name + ","  # col 4 母号
               + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
               + child_id + ','
               + popus[pre_pos].name + ","  # col 9 公号
               + popus[male_num + cur_female].name + ","  # col 10 母号
               + tmp_fid + ","
               + sex_id + '\n')  # 11 家系号
    best_idx += 1

    pre_pos = best_solution.vector_male[0]
    cur_female = best_solution.vector_female[0]
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    sex_id = "0"
    child_id = idgenarator.get_new_id()
    fout.write(","  # col 1
               + tmp_fid + ","  # col 2 家系号
               + popus[pre_pos].name + ","  # col 3 公号
               + popus[male_num + cur_female].name + ","  # col 4 母号
               + f"{kinship_matrix[pre_pos, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
               + child_id + ','
               + popus[pre_pos].name + ","  # col 9 公号
               + popus[male_num + cur_female].name + ","  # col 10 母号
               + tmp_fid + ","
               + sex_id + '\n')  # 11 家系号
    fi += 1
    idx = 1
    # threshold = 0.5
    res_data = []
    while idx < len(best_solution):
        cur_male = best_solution.vector_male[idx]
        # cur_male_name = popus[cur_male].name
        cur_female = best_solution.vector_female[idx]
        # cur_female_name = popus[male_num + cur_female].name
        if cur_male != pre_pos:
            bst_pos = best_solution.vector_male[female_list[best_idx]]
            # cur_male_name = popus[cur_male].name
            bst_female = best_solution.vector_female[female_list[best_idx]]
            mi += 1
            tmp_fid = idgenarator.get_family_id(y="", m=mi)
            sex_id = "1"
            child_id = idgenarator.get_new_id()
            fout.write(","  # col 1
                       + tmp_fid + ","  # col 2 家系号
                       + popus[bst_pos].name + ","  # col 3 公号
                       + popus[male_num + bst_female].name + ","  # col 4 母号
                       + f"{kinship_matrix[bst_pos, bst_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
                       + child_id + ','
                       + popus[bst_pos].name + ","  # col 9 公号
                       + popus[male_num + bst_female].name + ","  # col 10 母号
                       + tmp_fid + ","
                       + sex_id + '\n')  # 11 家系号
            best_idx += 1

        # ibc = kinship_matrix[cur_male, cur_female]
        tmp_fid = idgenarator.get_family_id(y="", m=mi)
        sex_id = "0"
        child_id = idgenarator.get_new_id()
        fout.write(","  # col 1
                   + tmp_fid + ","  # col 2 家系号
                   + popus[cur_male].name + ","  # col 3 公号
                   + popus[male_num + cur_female].name + ","  # col 4 母号
                   + f"{kinship_matrix[cur_male, cur_female]:.5f},,,"  # col 5  亲缘相关系数 6 7
                   + child_id + ','
                   + popus[cur_male].name + ","  # col 9 公号
                   + popus[male_num + cur_female].name + ","  # col 10 母号
                   + tmp_fid + ","
                   + sex_id + '\n')

        pre_pos = cur_male
        idx += 1
        fi += 1
    # print("]")
    fout.write('\n')
    fout.close()
    print(f"generate finished gene {gene_idx}")
    return res_data


def run_main_with_graph(file_path, result_file=None, configs=None):
    """
    等量留种方式
    :param file_path:
    :param gene_idx:
    :param result_file:
    :return:
    """
    gene_idx = configs["gene_idx"]
    idgenarator = IDGenerator(end_number=int(gene_idx) * 1000, year=str(int(gene_idx) - 1))
    layergraph, vertex_layer, vertex_list, sheet_list = get_graph_from_data(file_path=file_path)
    # if len(sheet_list[0]) == 2:
    #     for i in range(len(sheet_list)):
    #         sheet_list[i] = "20"+sheet_list[i]
    print(sheet_list)
    layergraph.print_children()
    # print(sheet_list)
    kinship = Kinship(graph=layergraph)
    sheet_list += [str(int(sheet_list[-1]) + 1)]
    # print("Sheet list:", sheet_list)
    year2idx = {}  # {"16": 0, "17": 1, "18": 2, "19": 3, "20": 4, "21": 5}
    for jdx, item in enumerate(sheet_list):
        year2idx[item] = jdx
    popus = []
    # print("Sheet list:", year2idx)
    # print("Load edges from", gene_idx)
    # print("year idx", year2idx[gene_idx])
    # return
    if gene_idx == sheet_list[-1]:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            # print(vertex_list[item].name)
            popus.append(vertex_list[item])
    else:
        # print(vertex_layer)
        for idx, item in enumerate(vertex_layer[year2idx[gene_idx]]):
            # print(vertex_list[item].name)
            popus.append(vertex_list[item])
            # popus.append(Poultry(fi=f_i, wi=wi, fa_i=fa_i, ma_i=ma_i, sex=0, inbreedc=0.))
    # return
    # print(kinship.calc_inbreed_coef(p="14774"))
    # kinship.print_edges()
    # kinship.print_layer()
    # kinship.print_parents()
    # kinship.print_all_poultry()
    # print(kinship.calc_kinship_corr(p1="14761", p2="14766"))
    # random.shuffle(popus)
    male_rate = 1. / 11.
    male_num = 0
    female_num = 0
    male_indices = []
    female_indices = []
    # print("given genders:")
    for i in range(len(popus)):
        # print(popus[i].gender)
        if popus[i].gender in ["公", "1"]:
            male_num += 1
            male_indices.append(i)
        elif popus[i].gender in ["母", "0"]:
            female_num += 1
            female_indices.append(i)
        else:
            # raise Exception("Gender Error.")
            if np.random.rand() < male_rate:
                male_num += 1
                male_indices.append(i)
            else:
                female_num += 1
                female_indices.append(i)
        # popus[i].sex = 1
    print("number:", male_num, female_num)
    print("len:", male_indices, female_indices)
    print([popus[j].name for j in male_indices])
    print([popus[j].name for j in female_indices])

    # return
    name2idx = dict()
    for i, p in enumerate(popus):
        name2idx[p.name] = i
    # -------------Kinship read and build-------------

    kinship_matrix = np.zeros((male_num, female_num))
    for i in range(male_num):
        for j in range(female_num):
            kinship_matrix[i, j] = kinship.calc_kinship_corr(p1=popus[male_indices[i]].name,
                                                             p2=popus[female_indices[j]].name)
    print(kinship_matrix)
    print("max min")
    print(np.max(kinship_matrix), np.min(kinship_matrix))
    print(np.sum(kinship_matrix))

    # ===============================Algorithm==================================
    # ====================Here is vanilla Genetic Algorithm=====================

    GAS = GASelector(popus=popus, kinship_matrix=kinship_matrix, male_idxs=male_indices,
                     female_idxs=female_indices, num_popu=300, num_iter=50)
    best_solution = GAS.scheduler(mode=configs["mode"])

    # ===========================找出最佳雌性，来生育最佳雄性，等数留种=====================
    print("best solution:")
    print(best_solution.vector_male)
    print(best_solution.vector_female)

    # =====-------------===============
    # return

    female_list = []
    # tmp_female_idx = 0
    i, j = 0, 0
    bst_female_value = 999.
    bst_female_idx = None
    cur_male_idx = best_solution.vector_male[0]
    while i < len(best_solution):
        tmp_male_idx = best_solution.vector_male[i]
        if cur_male_idx == tmp_male_idx:
            tmp_female_idx = best_solution.vector_female[i]
            if kinship_matrix[tmp_male_idx, tmp_female_idx] < bst_female_value:
                bst_female_value = kinship_matrix[tmp_male_idx, tmp_female_idx]
                bst_female_idx = tmp_female_idx
        else:
            female_list.append(bst_female_idx)
            bst_female_value = 999.
            cur_male_idx = tmp_male_idx
        i += 1
    female_list.append(bst_female_idx)
    print("best females:", len(female_list))
    print(female_list)
    # =============================================================================

    # pre_pos = best_solution.vector_male[0]
    # cur_female = best_solution.vector_female[0]
    # print("========----------育种方案----------==========")
    # print("(家系号，雄性个体编号, 雌性个体编号)]")
    # idx = 1
    # fout = open(result_file, 'w', encoding="utf_8")
    # fout.write(
    #     "配种方案,家系号,公鸡号,母鸡号,亲缘相关系数,出雏,批次,翅号,公鸡号,母鸡号,{}年家系号,性别\n".format(gene_idx))
    All_Data_List = []
    # print("配种方案", "家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号", "{}年家系号".format(gene_idx))

    mi, fi = 0, 0
    best_idx = 0
    pre_pos = best_solution.vector_male[0]
    cur_female = female_list[best_idx]
    male_pos = male_indices[pre_pos]
    female_pos = female_indices[cur_female]
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    child_id = idgenarator.get_new_id()
    All_Data_List.append([None, tmp_fid, popus[male_pos].name, popus[female_pos].name,
                          f"{kinship_matrix[pre_pos, cur_female]:.5f}", None, None, child_id, popus[male_pos].name,
                          popus[female_pos].name, tmp_fid, "公"])
    best_idx += 1

    pre_pos = best_solution.vector_male[0]
    cur_female = best_solution.vector_female[0]
    male_pos = male_indices[pre_pos]
    female_pos = female_indices[cur_female]
    tmp_fid = idgenarator.get_family_id(y="", m=mi)
    child_id = idgenarator.get_new_id()
    All_Data_List.append([None, tmp_fid, popus[male_pos].name, popus[female_pos].name,
                          f"{kinship_matrix[pre_pos, cur_female]:.5f}", None, None, child_id, popus[male_pos].name,
                          popus[female_pos].name, tmp_fid, "母"])
    fi += 1
    idx = 1

    while idx < len(best_solution):
        cur_male = best_solution.vector_male[idx]
        cur_female = best_solution.vector_female[idx]
        male_pos = male_indices[cur_male]
        female_pos = female_indices[cur_female]
        if cur_male != pre_pos:
            bst_male = best_solution.vector_male[idx]
            bst_female = female_list[best_idx]
            bst_male_pos = male_indices[bst_male]
            bst_female_pos = female_indices[bst_female]
            mi += 1
            tmp_fid = idgenarator.get_family_id(y="", m=mi)
            child_id = idgenarator.get_new_id()

            All_Data_List.append([None, tmp_fid, popus[bst_male_pos].name, popus[bst_female_pos].name,
                                  f"{kinship_matrix[bst_male, bst_female]:.5f}", None, None, child_id,
                                  popus[bst_male_pos].name,
                                  popus[bst_female_pos].name, tmp_fid, "公"])
            best_idx += 1

        # ibc = kinship_matrix[cur_male, cur_female]
        tmp_fid = idgenarator.get_family_id(y="", m=mi)
        child_id = idgenarator.get_new_id()
        All_Data_List.append([None, tmp_fid, popus[male_pos].name, popus[female_pos].name,
                              f"{kinship_matrix[cur_male, cur_female]:.5f}", None, None, child_id, popus[male_pos].name,
                              popus[female_pos].name, tmp_fid, "母"])
        pre_pos = cur_male
        idx += 1
        fi += 1

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(result_file.format(gene_idx), engine='openpyxl')
    writer.book = book
    # df1 = pd.DataFrame(np.array(res_data))
    df1 = pd.DataFrame(np.array(All_Data_List))
    df1.columns = ["配种方案", "家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号",
                   "{}年家系号".format(gene_idx), "性别"]
    df1.to_excel(writer, gene_idx, index=False)  # first是第一张工作表名称
    writer.save()
    writer.close()
    print(f"generate finished gene {gene_idx}")


# def run_main_without_graph_niter(file_path="./kinship330.csv", gene_idx=2018, result_file=None):
#     max_year = 2020
#     t_year = 2025
#     assert max_year + 1 < t_year + 1, "max_year:{} t_year:{}".format(max_year + 2, t_year + 1)
#     # df = pd.read_csv(file_path, delimiter=',', header=0, index_col=None)
#     # popus_idxs = list(df.iloc[:, 0]) + list(df.columns[1:])
#     # popus = []
#     # for idx, name in enumerate(popus_idxs):
#     #     popus.append(Vertex(idx, name=name))
#     # kinship_matrix = np.array(df.iloc[:, 1:])
#
#     res_data = None
#     for f_year in range(max_year + 1, t_year + 1):
#         # run_main(calc.file_to_analyze, gene_idx=str("f_year"))
#         print("open new csv file:", "./result_name_rand_{}.csv".format(f_year))
#         res_data = run_main_without_graph(file_path=file_path, gene_idx=f_year,
#                                           result_file=result_file.format(f_year))
#
#         book = load_workbook(cur_file)
#         writer = pd.ExcelWriter(name_template.format(f_year), engine='openpyxl')
#         writer.book = book
#         df1 = pd.DataFrame(np.array(res_data))
#         df1.columns = ["家系号", "公鸡号", "母鸡号", "亲缘相关系数", "出雏", "批次", "翅号", "公鸡号", "母鸡号",
#                        "{}年家系号".format(str(f_year)[-2:]), "性别"]
#         df1.to_excel(writer, str(f_year))  # first是第一张工作表名称
#         writer.save()
#         writer.close()
#         cur_file = name_template.format(f_year)
#     calc.generated_file = calc.file_root + cur_file


if __name__ == '__main__':
    # run_main_without_graph(file_path="./analyzer/kinship330.csv", result_file="./analyzer/output_{}.csv")
    run_main_with_graph(file_path="./analyzer/first330.xlsx", gene_idx=2025, result_file="./analyzer/output_{}.csv")
